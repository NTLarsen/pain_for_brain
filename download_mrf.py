import csv
from openpyxl import load_workbook
import clickhouse_driver
from time import sleep
import os
#pip3 install openpyxl

book = load_workbook('mrf_dict/mrf_dictionary.xlsx')
sheet = book.active

csv_data = []
for value in sheet.iter_rows(values_only=True):
    csv_data.append(list(value))

with open('mrf_dict.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(csv_data[1:])

user = "default"
password = ""
goals_table = "mrf_dict"
database = "db_metrica"

client = clickhouse_driver.Client(host='localhost', user=user, password=password)
client.execute("DROP TABLE IF EXISTS {}.{}".format(database, goals_table))
sleep(10)

client.execute('''
CREATE TABLE IF NOT EXISTS {}.{}
(
	"Region_name" String,
	"Host_type" String,
	"Mrf" String,
	"Regexp" String
)ENGINE = TinyLog;'''.format(database, goals_table))
sleep(10)

os.system('''cat mrf_dict.csv | clickhouse-client  --query="INSERT INTO {}.{} FORMAT CSV"'''.format(database, goals_table))

